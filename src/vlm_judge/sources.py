from __future__ import annotations

import hashlib
import json
import re
import statistics
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from .metrics import deterministic_match
from .schema import BenchmarkTask, EvaluationItem


QUESTION_TYPE_MAP = {
    "single-choice question": "multiple_choice",
    "multiple-choice": "multi_answer",
    "true-false": "short_text",
    "fill in the blanks": "multi_answer",
    "match the items": "multi_answer",
    "order the items": "multi_answer",
    "open question (precise answer)": "open_ended",
    "open question (arbitrary answer)": "open_ended",
}


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("source parsing requires the optional 'sources' dependencies") from exc
    return load_workbook


def _jsonable(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _row_dict(headers: list[str], row: Iterable[Any]) -> dict[str, Any]:
    return {
        headers[index]: _jsonable(value)
        for index, value in enumerate(row)
        if index < len(headers) and headers[index]
    }


def _asset_id(value: str, fallback: str) -> str:
    name = Path(urlparse(value).path).stem if value.startswith(("http://", "https://")) else value
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_")
    return safe or fallback


def _domain(value: str | None) -> str:
    if not value:
        return ""
    return urlparse(value).netloc.casefold()


def parse_main_workbook(path: Path) -> tuple[list[BenchmarkTask], dict[str, Any]]:
    load_workbook = _require_openpyxl()
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    tasks: list[BenchmarkTask] = []
    subject_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    question_type_counts: Counter[str] = Counter()
    format_counts: Counter[str] = Counter()
    source_domains: Counter[str] = Counter()
    all_row_subject_counts: Counter[str] = Counter()
    unresolved_assets = 0

    for row_number, row in enumerate(rows, start=2):
        value = _row_dict(headers, row)
        visual = str(value.get("Visual") or "").strip()
        reference = str(value.get("Correct answer") or "").strip()
        raw_subject = str(value.get("Subject") or "").strip()
        if raw_subject:
            all_row_subject_counts[raw_subject] += 1
        if not visual and not reference:
            continue
        task_id = _asset_id(visual, f"main_{row_number:04d}")
        is_url = visual.startswith(("http://", "https://"))
        if not is_url:
            unresolved_assets += 1
        question_type = str(value.get("Question type") or "").strip()
        raw_subject = raw_subject or "unknown"
        subject = raw_subject
        task = BenchmarkTask(
            task_id=task_id,
            subject=subject,
            grade=value.get("Class"),
            answer_type=QUESTION_TYPE_MAP.get(question_type, "unknown"),
            question_image_url=visual if is_url else None,
            reference_answer=reference or None,
            metadata={
                "source_workbook": path.name,
                "source_sheet": "Sheet1",
                "source_row": row_number,
                "source_url": value.get("Source"),
                "subject_raw": raw_subject,
                "question_asset_ref": None if is_url else visual,
                "record_type": value.get("Type"),
                "input_format": value.get("Input format"),
                "question_format": value.get("Question format"),
                "question_type_raw": question_type,
                "handwritten_answer_samples": value.get("handwriten answer samples"),
            },
        )
        task.validate(allow_unresolved_assets=True)
        tasks.append(task)
        subject_counts[subject] += 1
        type_counts[str(value.get("Type") or "unknown")] += 1
        question_type_counts[question_type or "unknown"] += 1
        format_counts[str(value.get("Question format") or "unknown")] += 1
        source_domains[_domain(value.get("Source")) or "unknown"] += 1

    planned_subjects: list[str] = []
    if "Sheet5" in workbook.sheetnames:
        in_plan = False
        for row in workbook["Sheet5"].iter_rows(values_only=True):
            first = row[0] if row else None
            second = row[1] if len(row) > 1 else None
            if first == "Grades" and second == "Subject":
                in_plan = True
                continue
            if not in_plan:
                continue
            if first == "Total" or second == "Total":
                break
            if isinstance(first, (int, float)) and second:
                planned_subjects.append(str(second).strip())

    inventory = {
        "records": len(tasks),
        "subjects": dict(subject_counts.most_common()),
        "task_subject_label_count": len(subject_counts),
        "raw_subject_labels_all_rows": dict(all_row_subject_counts.most_common()),
        "raw_subject_label_count_all_rows": len(all_row_subject_counts),
        "planned_subject_categories": planned_subjects,
        "planned_subject_category_count": len(planned_subjects),
        "record_types": dict(type_counts.most_common()),
        "question_types": dict(question_type_counts.most_common()),
        "question_formats": dict(format_counts.most_common()),
        "top_source_domains": dict(source_domains.most_common(30)),
        "unresolved_question_assets": unresolved_assets,
        "unique_task_ids": len({task.task_id for task in tasks}),
    }
    return tasks, inventory


def parse_math_workbook(path: Path) -> tuple[list[BenchmarkTask], dict[str, Any]]:
    load_workbook = _require_openpyxl()
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Sayfa1"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    tasks: list[BenchmarkTask] = []
    grade_counts: Counter[str] = Counter()
    difficulty_counts: Counter[str] = Counter()
    answer_type_counts: Counter[str] = Counter()
    topic_counts: Counter[str] = Counter()

    for row_number, row in enumerate(rows, start=2):
        value = _row_dict(headers, row)
        task_id = str(value.get("Question ID") or "").strip()
        if not task_id:
            continue
        flags = {
            "easy": bool(value.get("E")),
            "medium": bool(value.get("M")),
            "hard": bool(value.get("H")),
            "multiple_choice": bool(value.get("MC")),
            "open_ended_numeric": bool(value.get("OEN")),
            "symbolic": bool(value.get("SYM")),
            "multi_step": bool(value.get("MS")),
        }
        textual_answer = str(value.get("AMC") or "").strip()
        if ";" in textual_answer:
            answer_type = "multi_answer"
        elif flags["multiple_choice"]:
            answer_type = "multiple_choice"
        elif flags["open_ended_numeric"] and flags["multi_step"]:
            answer_type = "multi_answer"
        elif flags["open_ended_numeric"]:
            answer_type = "numeric"
        elif flags["symbolic"]:
            answer_type = "open_ended"
        else:
            answer_type = "unknown"

        task = BenchmarkTask(
            task_id=task_id,
            subject="Mathematics",
            grade=value.get("Grade"),
            answer_type=answer_type,
            question_image_url=str(value.get("URLQ") or "").strip() or None,
            reference_image_url=str(value.get("URLA") or "").strip() or None,
            reference_answer=textual_answer or None,
            metadata={
                "source_workbook": path.name,
                "source_sheet": "Sayfa1",
                "source_row": row_number,
                "answer_id": value.get("Answer ID"),
                "topic_area": value.get("Topic Area"),
                "sub_topic": value.get("Sub-topic"),
                "spec_labels": value.get("Spec Labels"),
                **flags,
            },
        )
        task.validate()
        tasks.append(task)
        grade_counts[str(value.get("Grade"))] += 1
        difficulty = next((name for name in ("easy", "medium", "hard") if flags[name]), "unlabeled")
        difficulty_counts[difficulty] += 1
        answer_type_counts[answer_type] += 1
        topic_counts[str(value.get("Topic Area") or "unknown")] += 1

    inventory = {
        "records": len(tasks),
        "unique_task_ids": len({task.task_id for task in tasks}),
        "grades": dict(grade_counts.most_common()),
        "difficulties": dict(difficulty_counts.most_common()),
        "answer_types": dict(answer_type_counts.most_common()),
        "top_topics": dict(topic_counts.most_common()),
        "question_urls_present": sum(bool(task.question_image_url) for task in tasks),
        "reference_image_urls_present": sum(bool(task.reference_image_url) for task in tasks),
        "text_references_present": sum(bool(task.reference_answer) for task in tasks),
    }
    return tasks, inventory


def parse_seed_failures(path: Path) -> tuple[list[EvaluationItem], list[dict[str, Any]]]:
    """Extract the eight historical wrong-answer examples from Sheet6."""
    load_workbook = _require_openpyxl()
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Sheet6"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    items: list[EvaluationItem] = []
    scores: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        value = _row_dict(headers, row)
        candidate = str(value.get("result") or "").strip()
        reference = str(value.get("answer") or "").strip()
        if not candidate or not reference:
            continue
        image_ref = str(value.get("image_url") or "").strip()
        item = EvaluationItem(
            task_id=f"legacy_failure_{row_number - 1:02d}",
            candidate_answer=candidate,
            subject=str(value.get("subject") or "unknown"),
            grade=value.get("class"),
            answer_type="multiple_choice",
            setup="unknown",
            question_image_url=image_ref or None,
            reference_answer=reference,
            metadata={
                "source_workbook": path.name,
                "source_sheet": "Sheet6",
                "source_row": row_number,
                "original_filename": value.get("image_filename"),
                "source_label_correct": value.get("is_result_correct"),
            },
        )
        item.validate()
        result = deterministic_match(reference, candidate, "multiple_choice")
        items.append(item)
        scores.append(
            {
                "task_id": item.task_id,
                "setup": item.setup,
                "subject": item.subject,
                "deterministic": {
                    "applicable": result.applicable,
                    "matched": result.matched,
                    "method": result.method,
                    "normalized_reference": result.normalized_reference,
                    "normalized_candidate": result.normalized_candidate,
                },
                "source_label_correct": value.get("is_result_correct"),
            }
        )
    return items, scores


def _percentile(sorted_values: list[int], quantile: float) -> int:
    index = min(len(sorted_values) - 1, int((len(sorted_values) - 1) * quantile))
    return sorted_values[index]


def inspect_odevjet(path: Path) -> dict[str, Any]:
    subject_counts: Counter[str] = Counter()
    grade_counts: Counter[str] = Counter()
    image_count_distribution: Counter[int] = Counter()
    publisher_counts: Counter[str] = Counter()
    missing_metadata: Counter[str] = Counter()
    books: set[Any] = set()
    page_urls: set[str] = set()
    lengths: list[int] = []
    seen_ids: dict[str, str] = {}
    duplicate_rows = 0
    conflicting_duplicate_ids: set[str] = set()
    low_information_records = 0
    parse_errors = 0
    records = 0

    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                parse_errors += 1
                continue
            records += 1
            metadata = item.get("metadata") or {}
            content = str(item.get("content") or "").strip()
            normalized = unicodedata.normalize("NFKC", content).casefold()
            if len(content) < 100 or "henüz çözüm" in normalized:
                low_information_records += 1
            lengths.append(len(content))
            item_id = str(item.get("id") or "")
            digest = hashlib.sha256(
                json.dumps(item, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest()
            if item_id in seen_ids:
                duplicate_rows += 1
                if seen_ids[item_id] != digest:
                    conflicting_duplicate_ids.add(item_id)
            else:
                seen_ids[item_id] = digest
            subject_counts[str(metadata.get("ders") or "unknown")] += 1
            grade_counts[str(metadata.get("sinif") or "unknown")] += 1
            publisher_counts[str(metadata.get("yayinevi") or "unknown")] += 1
            books.add(metadata.get("kitap_id"))
            if metadata.get("url"):
                page_urls.add(str(metadata["url"]))
            image_count_distribution[len(metadata.get("image_urls") or [])] += 1
            for key in ("kitap_id", "kitap_title", "sinif", "ders", "sayfa_no", "url", "image_urls"):
                if metadata.get(key) in (None, "", []):
                    missing_metadata[key] += 1

    lengths.sort()
    content_lengths = {}
    if lengths:
        content_lengths = {
            "min": min(lengths),
            "p25": _percentile(lengths, 0.25),
            "median": _percentile(lengths, 0.50),
            "p75": _percentile(lengths, 0.75),
            "p90": _percentile(lengths, 0.90),
            "p99": _percentile(lengths, 0.99),
            "max": max(lengths),
            "mean": round(statistics.mean(lengths), 1),
        }
    return {
        "records": records,
        "parse_errors": parse_errors,
        "unique_ids": len(seen_ids),
        "duplicate_rows": duplicate_rows,
        "conflicting_duplicate_ids": len(conflicting_duplicate_ids),
        "unique_books": len(books),
        "unique_page_urls": len(page_urls),
        "subjects": dict(subject_counts.most_common()),
        "grades": dict(grade_counts.most_common()),
        "top_publishers": dict(publisher_counts.most_common(20)),
        "image_count_distribution": {
            str(key): value for key, value in sorted(image_count_distribution.items())
        },
        "missing_metadata": dict(missing_metadata),
        "low_information_records": low_information_records,
        "content_length": content_lengths,
    }


def write_tasks(path: Path, tasks: Iterable[BenchmarkTask]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for task in tasks:
            handle.write(json.dumps(task.to_dict(), ensure_ascii=False) + "\n")


def write_records(path: Path, records: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def prepare_sources(
    main_workbook: Path,
    math_workbook: Path,
    corpus: Path,
    output_dir: Path,
) -> dict[str, Any]:
    main_tasks, main_inventory = parse_main_workbook(main_workbook)
    math_tasks, math_inventory = parse_math_workbook(math_workbook)
    seed_items, seed_scores = parse_seed_failures(main_workbook)
    corpus_inventory = inspect_odevjet(corpus)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_tasks(output_dir / "main_benchmark.jsonl", main_tasks)
    write_tasks(output_dir / "math_benchmark.jsonl", math_tasks)
    write_records(output_dir / "seed_failure_candidates.jsonl", (item.to_dict() for item in seed_items))
    write_records(output_dir / "seed_failure_scores.jsonl", seed_scores)
    inventory = {
        "main_workbook": main_inventory,
        "math_workbook": math_inventory,
        "seed_failure_validation": {
            "records": len(seed_scores),
            "deterministic_labels_matching_source": sum(
                (score["deterministic"]["matched"] is False)
                == (str(score["source_label_correct"]).casefold() == "no")
                for score in seed_scores
            ),
        },
        "odevjet_corpus": corpus_inventory,
        "known_source_mismatches": {
            "mentor_subject_count": 17,
            "workbook_planned_subject_categories": main_inventory["planned_subject_category_count"],
            "main_workbook_raw_subject_labels_all_rows": main_inventory["raw_subject_label_count_all_rows"],
            "main_workbook_task_subject_labels": main_inventory["task_subject_label_count"],
            "subject_mapping_status": "unconfirmed_keep_raw_labels",
        },
    }
    with (output_dir / "source_inventory.json").open("w", encoding="utf-8") as handle:
        json.dump(inventory, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    return inventory
